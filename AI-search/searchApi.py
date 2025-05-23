from flask import Flask, request, jsonify
import openpyxl
import os
from sentence_transformers import SentenceTransformer
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import pandas as pd
import joblib
from joblib import dump
from joblib import load
import urllib.parse

app = Flask(__name__)

model = SentenceTransformer('all-MiniLM-L6-v2')

clf = LogisticRegression()

EXCEL_FILE = "MatchWords.xlsx"


# Create Excel file with headers if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["String 1", "String 2"])
    wb.save(EXCEL_FILE)

@app.route("/addPair", methods=["POST"])
def add_to_excel():
    data = request.args
    string1 = data.get("string1")
    string2 = data.get("string2")
    string1 = string1.replace("%2520", " ")
    string2 = string2.replace("%2520", " ")
    acc =int(""+ data.get("same"))
    if not string1 or not string2:
        return jsonify({"error": "Both string1 and string2 are required."}), 400

    # Append data to Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([string1, string2,acc])
    wb.save(EXCEL_FILE)
    ACol =[]
    BCol = []
    labels =[]
    for row in ws.iter_rows( max_col=3, values_only=True):
        a,b,c = row
        ACol.append(a)
        BCol.append(b)
        labels.append(c==1)
    wb.close()
    emb1 = model.encode(ACol)
    emb2 = model.encode(BCol)
    
    
    features = np.abs(emb1 - emb2)
    

    # Train-test split
    X_train, X_test, y_train, y_test = train_test_split(features, labels, test_size=0.2)
    # print(X_train)
    # print(X_test)
    # print(y_test)
    # print(y_train)

    clf.fit(X_train, y_train)
    y_pred = clf.predict(X_test)
    print(labels)
    joblib.dump(clf, 'product_match_model.pkl')
    return jsonify("Accuracy:", accuracy_score(y_test, y_pred)), 200



@app.route("/checkPair", methods=["POST"])
def checkPair():
    data = request.args
    string1 = data.get("input")
    string2 = data.get("actualName")
    decoded_once = urllib.parse.unquote(string1)       # Decodes %25XX to %XX
    string1 = urllib.parse.unquote(decoded_once)
    decoded_once = urllib.parse.unquote(string2)       # Decodes %25XX to %XX
    string2 = urllib.parse.unquote(decoded_once)
    if not string1 or not string2:
        return jsonify({"error": "Both string1 and string2 are required."}), 400
    if string2 in string1:
        return {"match": bool(1), "confidence": 1}
    # wb = openpyxl.load_workbook(EXCEL_FILE)
    # ws = wb.active
    clf1 = joblib.load("product_match_model.pkl")
    emb1 = model.encode(string1)
    emb2 = model.encode(string2)
    features = np.abs(emb1 - emb2)
    prediction = clf1.predict([features])[0]
    confidence = clf1.predict_proba([features])[0][1]
    print("check pair "+string1+" and "+ string2 +". same percentt is "+str(round(float(confidence), 2)))
    return {"match": bool(prediction>0.8), "confidence": round(float(confidence), 2)}




if __name__ == "__main__":
    app.run(debug=True)
