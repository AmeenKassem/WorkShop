import openpyxl
from flask import Flask, request, jsonify
import os
from sentence_transformers import SentenceTransformer
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import pandas as pd
import joblib
from joblib import dump
from joblib import load
import urllib.parse

app = Flask(__name__)

model = SentenceTransformer('all-MiniLM-L6-v2')

clf = LogisticRegression()

EXCEL_FILE = "AI-Search-API\\MatchWords.xlsx"

import pyodbc

# Define connection parameters
server = 'ws-server.database.windows.net'
database = 'rahaf_db'
username = 'rahaf'
password = 'RaPass2025'
driver = '{ODBC Driver 17 for SQL Server}'  # Ensure this driver is installed

# Create the connection string
conn_str = f"""
    DRIVER={driver};
    SERVER={server};
    DATABASE={database};
    UID={username};
    PWD={password};
    Encrypt=yes;
    TrustServerCertificate=no;
    Connection Timeout=30;
"""

print("✅ Connected to Azure SQL Server.")



# Create Excel file with headers if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["String 1", "String 2"])
    wb.save(EXCEL_FILE)


@app.route("/get-matches-products",methods=["POST"])
def getMatchesProducts():
    data = request.args
    query = "SELECT * FROM product"
    categoryNum = int(data.get('category'))
    same = float(data.get('same'))
    if(categoryNum!=-1):
        query = query + " WHERE category="+categoryNum
    input = data.get("input")
    decoded_once = urllib.parse.unquote(input)       # Decodes %25XX to %XX
    input = urllib.parse.unquote(decoded_once)
    emb2 = model.encode(input)
    conn = pyodbc.connect(conn_str)
    df = pd.read_sql(query, conn)
    conn.close()
    print(df)
    productsResult = []
    clf1 = joblib.load("AI-Search-API\\product_match_model.pkl")
    for index,row in df.iterrows():
        productName =str(row['name'])
        decoded_once = urllib.parse.unquote(productName)       # Decodes %25XX to %XX
        productName = urllib.parse.unquote(decoded_once)
        emb1 = model.encode(productName)
        features = np.abs(emb1 - emb2)
        prediction = clf1.predict([features])[0]
        confidence = clf1.predict_proba([features])[0][1]
        print("check pair "+input+" and "+ productName +". same percentt is "+str(round(float(confidence), 2)))
        if bool(float(confidence)>same):
            append_sorted_by_confidence(productsResult,({'name':str(row['name']),'productId':int(row['product_id']),"confidence": round(float(confidence), 2)}))
        

    return productsResult, 200

def append_sorted_by_confidence(result_list, new_item):
    """
    Appends `new_item` to `result_list`, keeping it sorted by 'confidence' descending.
    """
    index = 0
    while index < len(result_list) and result_list[index]['confidence'] >= new_item['confidence']:
        index += 1
    result_list.insert(index, new_item)

@app.route("/addPair", methods=["POST"])
def add_to_excel():
    data = request.args
    string1 = data.get("string1")
    string2 = data.get("string2")
    string1 = string1.replace("%2520", " ")
    string2 = string2.replace("%2520", " ")
    acc =int(""+ data.get("same"))
    if not string1 or not string2:
        return jsonify({"error": "Both string1 and string2 are required."}), 400

    # Append data to Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([string1, string2,acc])
    wb.save(EXCEL_FILE)
    ACol =[]
    BCol = []
    labels =[]
    for row in ws.iter_rows( max_col=3, values_only=True):
        a,b,c = row
        ACol.append(a)
        BCol.append(b)
        labels.append(c==1)
    wb.close()
    emb1 = model.encode(ACol)
    emb2 = model.encode(BCol)
    
    
    features = np.abs(emb1 - emb2)
    

    # Train-test split
    X_train, X_test, y_train, y_test = train_test_split(features, labels, test_size=0.2)
    # print(X_train)
    # print(X_test)
    # print(y_test)
    # print(y_train)

    clf.fit(X_train, y_train)
    y_pred = clf.predict(X_test)
    print(labels)
    joblib.dump(clf, 'AI-Search-API\\product_match_model.pkl')
    return jsonify("Accuracy:", accuracy_score(y_test, y_pred)), 200



@app.route("/checkPair", methods=["POST"])
def checkPair():
    data = request.args
    string1 = data.get("input")
    string2 = data.get("actualName")
    decoded_once = urllib.parse.unquote(string1)       # Decodes %25XX to %XX
    string1 = urllib.parse.unquote(decoded_once)
    decoded_once = urllib.parse.unquote(string2)       # Decodes %25XX to %XX
    string2 = urllib.parse.unquote(decoded_once)
    if not string1 or not string2:
        return jsonify({"error": "Both string1 and string2 are required."}), 400
    if string2 in string1:
        return {"match": bool(1), "confidence": 1}
    # wb = openpyxl.load_workbook(EXCEL_FILE)
    # ws = wb.active
    clf1 = joblib.load("AI-Search-API\\product_match_model.pkl")
    emb1 = model.encode(string1)
    emb2 = model.encode(string2)
    features = np.abs(emb1 - emb2)
    prediction = clf1.predict([features])[0]
    confidence = clf1.predict_proba([features])[0][1]
    print("check pair "+string1+" and "+ string2 +". same percentt is "+str(round(float(confidence), 2)))
    return {"match": bool(float(confidence)>0.3), "confidence": round(float(confidence), 2)}


@app.route("/addPairs", methods=["POST"])
def add_multiple_to_excel():
    data = request.get_json()

    string1 = data.get("string1")
    string2_list = data.get("string2_list")  # Expecting a list
    acc = int(data.get("same", 0))

    if not string1 or not string2_list or not isinstance(string2_list, list):
        return jsonify({"error": "string1 and string2_list (as array) are required."}), 400

    string1 = string1.replace("%2520", " ")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Append data to Excel
    for string2 in string2_list:
        string2 = string2.replace("%2520", " ")
        ws.append([string1, string2, acc])

    wb.save(EXCEL_FILE)

    # Rebuild training data
    ACol, BCol, labels = [], [], []
    for row in ws.iter_rows(max_col=3, values_only=True):
        a, b, c = row
        if a is None or b is None or c is None:
            continue
        ACol.append(a)
        BCol.append(b)
        labels.append(int(c))  # Use 0/1 directly instead of True/False

    wb.close()

    for row in ws.iter_rows( max_col=3, values_only=True):
        a,b,c = row
        ACol.append(a)
        BCol.append(b)
        labels.append(c==1)
    wb.close()
    emb1 = model.encode(ACol)
    emb2 = model.encode(BCol)
    
    
    features = np.abs(emb1 - emb2)
    

    # Train-test split
    X_train, X_test, y_train, y_test = train_test_split(features, labels, test_size=0.2)
    # print(X_train)
    # print(X_test)
    # print(y_test)
    # print(y_train)

    clf.fit(X_train, y_train)
    y_pred = clf.predict(X_test)
    print(labels)
    joblib.dump(clf, 'AI-Search-API\\product_match_model.pkl')
    return jsonify("Accuracy:", accuracy_score(y_test, y_pred)), 200

    # Check if we have at least 2 classes
    if len(set(labels)) < 2:
        return jsonify({"error": "Model needs at least 2 classes (0 and 1) to train."}), 400

    emb1 = model.encode(ACol)
    emb2 = model.encode(BCol)
    features = np.abs(emb1 - emb2)

    # Train model
    X_train, X_test, y_train, y_test = train_test_split(features, labels, test_size=0.2)
    clf.fit(X_train, y_train)
    y_pred = clf.predict(X_test)

    joblib.dump(clf, 'product_match_model.pkl')

    return jsonify({
        "message": f"Added {len(string2_list)} pairs.",
        "accuracy": accuracy_score(y_test, y_pred)
    }), 200




if __name__ == "__main__":
    app.run(debug=True)
